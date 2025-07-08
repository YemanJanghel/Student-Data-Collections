from tkinter import*
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

background="#06283D"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Student Registration")
root.geometry("1250x700")

root.config(bg=background)

file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date Of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's occupation"
    sheet['L1'] = "Mother's occupation"

    file.save('Student_data.xlsx')


######################################################################333exit
def exit():
    root.destroy()


###################################################################upload
def upload():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetype=(("JPG File","*.jpg"),
                                                                            ("PNG File","*.png"),
                                                                          ("All files","*.txt")))
    img=(Image.open(filename))
    resized=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized)
    lbl.config(image=photo2)
    lbl.image=photo2


##################################################################registration
def registration():
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Registation.set(max_row_value+1)

    except:
        Registation.set("1")


###################################################clear
def clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    Father.set('')
    Mother.set('')
    Occupation1.set('')
    Occupation2.set('')
    Class.set("Select Class")

    registration()

    save.config(state="normal")

    img1=PhotoImage(file="Images/upload photo.png")
    lbl.config(image=img1)
    lbl.image=img1

    img=""


#####################################################################Save
def save():
    R1=Registation.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","select Gender!!!!!!")
    D1=DOB.get()
    D2=Date.get()
    Rel=Religion.get()
    S1=Skill.get()
    F1=father.get()
    M1=mother.get()
    F2=occupation1.get()
    M2=occupation2.get()
    if N1=="" or C1=="Select Class" or D1=="" or D2=="" or Rel=="" or S1=="" or F1=="" or M1=="" or F2=="" or M2=="":
        messagebox.showerror("error","Few Data is Missing!!!")
    else:
        file=openpyxl.load_workbook("Student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2, row=sheet.max_row , value=N1)
        sheet.cell(column=3, row=sheet.max_row , value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row , value=D1)
        sheet.cell(column=6, row=sheet.max_row , value=D2)
        sheet.cell(column=7, row=sheet.max_row , value=Rel)
        sheet.cell(column=8, row=sheet.max_row , value=S1)
        sheet.cell(column=9, row=sheet.max_row , value=F1)
        sheet.cell(column=10, row=sheet.max_row , value=M1)
        sheet.cell(column=11, row=sheet.max_row, value=F2)
        sheet.cell(column=12, row=sheet.max_row , value=M2)

        file.save(r"Student_data.xlsx")

        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!!!")

        messagebox.showinfo("info","Successfully data entered!!!!")
        clear()
        registration()


################################################################33search
def search():
    text=Search.get()
    clear()
    save.config(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            #print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number = str(name)[15:-1]



    try:
        print(str(name))
    except:
        messagebox.showinfo("Invalid","Invalid registration number!!!!!")


    x1 = sheet.cell(row=int(reg_number),column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    Registation.set(x1)
    Name.set(x2)
    Class.set(x3)
    if x4 == 'Female':
        r2.select()
    else:
        r1.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    Father.set(x9)
    Mother.set(x10)
    Occupation1.set(x11)
    Occupation2.set(x12)

    img=(Image.open("Student Images/"+str(x1)+".jpg"))
    resized = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized)
    lbl.config(image=photo2)
    lbl.image = photo2


###########################################3update
def update():
    R1 = Registation.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1=gender
    D1 = DOB.get()
    D2 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    F1 = father.get()
    M1 = mother.get()
    F2 = occupation1.get()
    M2 = occupation2.get()

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            print(reg_number)

    #sheet.cell(column=1, row=int(reg_number), value=R1)
    sheet.cell(column=2, row=int(reg_number), value=N1)
    sheet.cell(column=3, row=int(reg_number), value=C1)
    sheet.cell(column=4, row=int(reg_number), value=G1)
    sheet.cell(column=5, row=int(reg_number), value=D1)
    sheet.cell(column=6, row=int(reg_number), value=D2)
    sheet.cell(column=7, row=int(reg_number), value=Rel)
    sheet.cell(column=8, row=int(reg_number), value=S1)
    sheet.cell(column=9, row=int(reg_number), value=F1)
    sheet.cell(column=10, row=int(reg_number), value=M1)
    sheet.cell(column=11, row=int(reg_number), value=F2)
    sheet.cell(column=12, row=int(reg_number), value=M2)

    file.save(r"Student_data.xlsx")

    try:
        img.save("Student Images/"+str(R1)+".jpg")
    except:
        pass

    messagebox.showinfo("Update","Update Successfully!!")

    clear()







#$#########################################################gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"


###################################################top frame
Label(root,text="Email: Yemanjanghel555@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg="#fff",font="arial 20 bold").pack(side=TOP,fill=X)

########################################################searching box
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
imageicon3=PhotoImage(file="Images/search.png")
srch=Button(root,text="SEARCH",compound=LEFT,image=imageicon3,width=123,bg='sky blue',font='arial 13 bold',command=search)
srch.place(x=1060,y=66)

imageicon4=PhotoImage(file="Images/Layer 4.png")
update=Button(root,image=imageicon4,bg="#c36464",command=update)
update.place(x=110,y=64)


#registration and date
Label(root,text="Registration No :",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Registation=IntVar()
Date=StringVar()

reg=Entry(root,textvariable=Registation,width=15,font="arial 10")
reg.place(x=160,y=152)

registration()

today=date.today()
d1=today.strftime("%d/%m/%Y")
date=Entry(root,textvariable=Date,width=15,font="arial 10")
date.place(x=550,y=150)
Date.set(d1)


#Student details
obj=LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)
Label(obj,text="Full Name :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date Of Birth :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name=Entry(obj,textvariable=Name,width=20,font="arial 10")
name.place(x=160,y=53)

DOB=StringVar()
dob=Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob.place(x=160,y=103)

radio=IntVar()
r1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
r1.place(x=150,y=150)
r2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
r2.place(x=200,y=150)

Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 10",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")

Religion=StringVar()
religion=Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion.place(x=630,y=103)

Skill=StringVar()
skill=Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill.place(x=630,y=153)




#parent details
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)
Label(obj2,text="Father's Name :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

Father=StringVar()
father=Entry(obj2,textvariable=Father,width=20,font="arial 10")
father.place(x=160,y=53)

Occupation1=StringVar()
occupation1=Entry(obj2,textvariable=Occupation1,width=20,font="arial 10")
occupation1.place(x=160,y=100)


Label(obj2,text="Mother's Name :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

Mother=StringVar()
mother=Entry(obj2,textvariable=Mother,width=20,font="arial 10")
mother.place(x=630,y=53)

Occupation2=StringVar()
occupation2=Entry(obj2,textvariable=Occupation2,width=20,font="arial 10")
occupation2.place(x=630,y=103)


#images
f=Frame(root,bd=3,bg="black",width=201,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="Images/upload photo.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)


#button
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=upload).place(x=1000,y=370)
save=(Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="green",command=save))
save.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="orange",command=clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="red",command=exit).place(x=1000,y=630)



root.mainloop()